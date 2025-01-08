from datetime import datetime, time
from django.utils import timezone
from django.shortcuts import render, get_object_or_404, redirect
from .models import Address, EventAddress, Project, Event, RequestCounter, EventUser
from .forms import ProjectForm, EventForm
from django.contrib.auth.decorators import login_required
from openpyxl import load_workbook
import requests, json
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.utils.html import escape
from .utils import reset_request_counter, get_time_until_midnight, has_photos
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User, Group
from scipy.spatial.distance import euclidean
from django.utils.timezone import make_aware
from django.db.models import Max
import os
import re
import zipfile
from django.core.files.storage import default_storage
from django.conf import settings
from django.utils.encoding import iri_to_uri

@login_required
def manage_projects(request):
    # Получаем все проекты пользователя
    projects = Project.objects.filter(user=request.user)

    # Отделяем проекты по состоянию их событий, добавляя при этом ближайшее событие
    active_projects = []
    completed_projects = []

    for project in projects:
        # Получаем ближайшее событие для проекта
        next_event = project.events.filter(event_date__gte=timezone.now()).order_by('event_date').first()

        # Проверяем, имеют ли все события статус 'completed'
        if all(event.status == 'completed' for event in project.events.all()):
            completed_projects.append((project, next_event))
        else:
            active_projects.append((project, next_event))

    # Сортируем проекты по дате ближайшего события
    active_projects.sort(key=lambda x: x[1].event_date if x[1] else timezone.datetime.max)
    completed_projects.sort(key=lambda x: x[1].event_date if x[1] else timezone.datetime.max)

    return render(request, 'projects/manage_projects.html', {
        'active_projects': active_projects,
        'completed_projects': completed_projects
    })

@login_required
def edit_project(request, project_id):
    project = get_object_or_404(Project, pk=project_id)
    events = project.events.all()
    addresses = project.addresses.all()
    counter, created = RequestCounter.objects.get_or_create(user=request.user)
    remaining_requests = counter.count

    if request.method == 'POST':
        form = ProjectForm(request.POST, request.FILES, instance=project)
        if form.is_valid():
            if form.cleaned_data['organization'] == 'add_new':
                new_org = form.cleaned_data['new_organization']
                form.instance.organization = new_org
            else:
                form.instance.organization = form.cleaned_data['organization']
            # form.instance.manager остается без изменений, если это вдруг не предусмотрено в форме
            form.save()
            return redirect('edit_project', project_id=project.id)
    else:
        form = ProjectForm(instance=project)

    return render(request, 'projects/edit_project.html', {
        'form': form,
        'project': project,
        'events': events,
        'addresses': addresses,
        'remaining_requests': remaining_requests,
    })

@login_required
def create_project(request):
    if request.method == 'POST':
        form = ProjectForm(request.POST, request.FILES)
        if form.is_valid():
            project = form.save(commit=False)
            if form.cleaned_data['organization'] == 'add_new':
                new_org = form.cleaned_data['new_organization']
                project.organization = new_org
            else:
                project.organization = form.cleaned_data['organization']
            project.user = request.user
            project.save()
            return redirect('edit_project', project_id=project.id)
    else:
        form = ProjectForm()

    return render(request, 'projects/create_project.html', {'form': form})

def manage_events(request, project_id):
    project = get_object_or_404(Project, pk=project_id)
    events = project.events.all()
    return render(request, 'projects/manage_events.html', {'project': project, 'events': events})

def edit_event(request, event_id):
    event = get_object_or_404(Event, pk=event_id)
    project = event.project

    if request.method == 'POST':
        form = EventForm(request.POST, instance=event)
        if form.is_valid():
            form.save()
            event.update_status()  # Обновление статуса события.
            return redirect('edit_event', event_id=event.id)
    else:
        form = EventForm(instance=event)

    project_addresses = project.addresses.all()
    executor_group = Group.objects.get(name='Исполнитель')
    executors = User.objects.filter(groups=executor_group).prefetch_related('executorprofile')
    event_users = event.eventuser_set.select_related('user')
    user_profile = request.user.profile
    api_key = user_profile.api_key if user_profile else None

    # Проверяем есть ли неназначенные адреса
    has_unassigned_addresses = event.addresses.filter(assigned_user__isnull=True).exists()
    executors_with_photo_status = []

    for event_user in event_users:
        has_photos_status = has_photos(event_user.user, event)
        executors_with_photo_status.append((event_user, has_photos_status))

    event.update_status()  # Обновление статуса события.

    return render(request, 'projects/edit_event.html', {
        'form': form,
        'project': project,
        'event': event,
        'project_addresses': project_addresses,
        'executors': executors,
        'executors_with_photo_status': executors_with_photo_status,
        'api_key': api_key,
        'has_unassigned_addresses': has_unassigned_addresses,
    })

# перестраивает порядок адресов для оптимального маршрута
def nearest_neighbor(coords):
    n = len(coords)
    if n == 0:
        return []  # Возвращаем пустой путь, если нет координат

    visited = [False] * n
    path = [0]  # Начинаем с первой точки
    visited[0] = True

    for _ in range(n - 1):
        last_index = path[-1]
        nearest_index = -1
        nearest_distance = float('inf')

        for i in range(n):
            if not visited[i]:
                distance = euclidean(coords[last_index], coords[i])
                if distance < nearest_distance:
                    nearest_distance = distance
                    nearest_index = i

        path.append(nearest_index)
        visited[nearest_index] = True

    return path

# перестраивает порядок адресов для оптимального маршрута для деталей события
def nearest_neighbor_event_detail(data):
    def parse_coordinate(coord):
        # Заменяем запятую на точку и преобразуем строку в float
        return float(coord.replace(',', '.'))

    coords = [(parse_coordinate(item['latitude']), parse_coordinate(item['longitude'])) for item in data]
    num_points = len(coords)

    # Проверка формата координат
    for coord in coords:
        if len(coord) != 2:
            raise ValueError("Each coordinate must contain exactly two elements: latitude and longitude.")

    # Найти ближайшего соседа
    optimal_route = []
    visited = [False] * num_points
    current_index = 0
    for _ in range(num_points):
        visited[current_index] = True
        optimal_route.append(current_index)

        # Следующая ближайшая точка
        min_distance = float('inf')
        next_index = None
        for i in range(num_points):
            if not visited[i]:
                try:
                    distance = euclidean(coords[current_index], coords[i])
                    if distance < min_distance:
                        min_distance = distance
                        next_index = i
                except ValueError as e:
                    print("Caught exception:", e)
                    print(f"Current index: {current_index}, Next index: {i}")
                    print(f"Current coord: {coords[current_index]}, Next coord: {coords[i]}")

        if next_index is None:
            break

        current_index = next_index

    return optimal_route

# копирование адресов с проекта в событие с построением маршрута начиная с 1 адреса в проекте
def copy_addresses(request, event_id):
    event = get_object_or_404(Event, id=event_id)

    # Получаем адреса из проекта
    project_addresses = list(event.project.addresses.all())

    # Проверяем наличие адресов в проекте
    if not project_addresses:
        return JsonResponse({"error": "Нету адресов для синхронизации. Добавьте адреса в проект!"}, status=400)

    # Удаляем все существующие адреса для события
    event.addresses.all().delete()

    # Подготавливаем координаты для оптимизации
    coordinates = [(addr.latitude, addr.longitude) for addr in project_addresses]
    
    # Находим оптимальный порядок адресов только если есть координаты
    if coordinates:
        optimal_order = nearest_neighbor(coordinates)
        sorted_project_addresses = [project_addresses[i] for i in optimal_order]

        # Копируем адреса из проекта в событие в оптимальном порядке
        for addr in sorted_project_addresses:
            EventAddress.objects.create(
                event=event,
                name=addr.name,
                latitude=addr.latitude,
                longitude=addr.longitude,
            )

    # Подготавливаем данные для возврата через JSON
    event_addresses = list(event.addresses.values('name', 'latitude', 'longitude'))
    return JsonResponse({"message": "Addresses copied successfully", "addresses": event_addresses}, status=200)

def create_event(request, project_id):
    project = get_object_or_404(Project, pk=project_id)
    if request.method == 'POST':
        form = EventForm(request.POST)
        if form.is_valid():
            event = form.save(commit=False)
            event.project = project
            event.save()
            return redirect('edit_event', event_id=event.id)
    else:
        form = EventForm()
    return render(request, 'projects/create_event.html', {'form': form, 'project': project})

def delete_project(request, project_id):
    project = get_object_or_404(Project, pk=project_id)
    project.delete()
    return redirect('manage_projects')

def delete_event(request, event_id):
    event = get_object_or_404(Event, pk=event_id)
    project_id = event.project.id
    event.delete()
    return redirect('edit_project', project_id=project_id)

@login_required
def get_coordinates_from_yandex(request, address):
    profile = request.user.profile
    api_key = profile.api_key

    try:
        response = requests.get("https://geocode-maps.yandex.ru/1.x/", params={
            'apikey': api_key,
            'geocode': address,
            'format': 'json',
        })
        
        response_data = response.json()
        # Проверяем наличие данных в ответе
        if 'featureMember' not in response_data['response']['GeoObjectCollection'] or not response_data['response']['GeoObjectCollection']['featureMember']:
            raise ValueError("Invalid API Key or Geocode failed")
        
        geo_object = response_data['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']
        coordinates = geo_object['Point']['pos'].split()
        return {
            'lat': float(coordinates[1]),
            'lon': float(coordinates[0]),
        }
    
    except Exception as e:
        raise ValueError("Invalid API Key or Geocode failed")

@login_required
def get_coordinates(request, project_id):
    reset_request_counter(request)  # Обновляем счётчик запросов
    counter = RequestCounter.objects.get(user=request.user)

    if counter.count <= 0:
        time_left = get_time_until_midnight()
        return JsonResponse({
            'status': 'error',
            'message': f'Закончились запросы для координат - запросы будут доступны через {time_left}',
            'remaining_requests': counter.count
        })
    
    profile = request.user.profile
    if not profile.api_key:
        return JsonResponse({'status': 'error', 'message': 'Вам нужно заполнить поле API Key в профиле'})

    if request.method == 'POST' and request.FILES.get('excel_file'):
        project = get_object_or_404(Project, pk=project_id)
        uploaded_file = request.FILES['excel_file']
        workbook = load_workbook(uploaded_file, data_only=True)
        sheet = workbook.active
        addresses_list = []

        try:
            num_processed = 0

            for index, row in enumerate(sheet.iter_rows(min_row=2, max_col=1, values_only=True)):
                if row[0]:
                    if counter.count <= 0:
                        time_left = get_time_until_midnight()
                        return JsonResponse({
                            'status': 'error',
                            'message': f'Закончились запросы для координат - запросы будут доступны через {time_left}',
                            'remaining_requests': counter.count
                        })

                    try:
                        coordinates = get_coordinates_from_yandex(request, row[0])
                        address = Address.objects.create(
                            project=project,
                            name=row[0],
                            latitude=coordinates['lat'],
                            longitude=coordinates['lon']
                        )
                        html = f"""
                        <li data-id="{address.id}">
                            <span class="address-number">{index + 1}.</span>
                            <input class="address-name" value="{escape(address.name)}" data-id="{address.id}"> |
                            <span class="latitude">{address.latitude}</span> |
                            <span class="longitude">{address.longitude}</span>
                            <button class="delete-address-btn" data-id="{address.id}">Удалить</button>
                        </li>
                        """
                        addresses_list.append(html)
                        num_processed += 1
                    except Exception as e:
                        return JsonResponse({'status': 'error', 'message': 'Некорректный API Key', 'remaining_requests': counter.count})
            
            counter.count -= num_processed
            counter.save()

            return JsonResponse({'status': 'ok', 'addresses': addresses_list, 'remaining_requests': counter.count})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e), 'remaining_requests': counter.count})

    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

@require_POST
def delete_addresses(request, project_id):
    project = get_object_or_404(Project, pk=project_id)
    # Удаляем все адреса, связанные с проектом
    project.addresses.all().delete()
    return JsonResponse({'status': 'ok'})

@require_http_methods(["POST"])
def edit_address_name(request, project_id):
    payload = json.loads(request.body)
    address_id = payload.get('id')
    new_name = payload.get('new_name')

    try:
        address = Address.objects.get(id=address_id, project_id=project_id)
        address.name = new_name
        address.save()
        return JsonResponse({'status': 'ok'})
    except Address.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Address not found'}, status=404)

@require_http_methods(["DELETE"])
def delete_address(request, project_id):
    payload = json.loads(request.body)
    address_id = payload.get('id')

    try:
        address = Address.objects.get(id=address_id, project_id=project_id)
        address.delete()
        return JsonResponse({'status': 'ok'})
    except Address.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Address not found'}, status=404)
    
@login_required
@require_POST
def update_addresses(request, project_id):
    reset_request_counter(request)
    counter = RequestCounter.objects.get(user=request.user)

    if counter.count <= 0:
        time_left = get_time_until_midnight()
        return JsonResponse({
            'status': 'error',
            'message': f'Закончились запросы для координат - запросы будут доступны через {time_left}',
            'remaining_requests': counter.count
        })

    profile = request.user.profile
    if not profile.api_key:
        return JsonResponse({'status': 'error', 'message': 'Вам нужно заполнить поле API Key в профиле'})

    try:
        project = get_object_or_404(Project, pk=project_id)
        data = json.loads(request.body)
        new_addresses = data.get('new_addresses', [])
        num_processed = 0

        max_order = Address.objects.filter(project=project).aggregate(Max('order'))['order__max'] or 0

        for name in new_addresses:
            if name:
                if counter.count <= 0:
                    time_left = get_time_until_midnight()
                    return JsonResponse({
                        'status': 'error',
                        'message': f'Закончились запросы для координат - запросы будут доступны через {time_left}',
                        'remaining_requests': counter.count
                    })
                
                try:
                    coordinates = get_coordinates_from_yandex(request, name)
                    new_order = max_order + 1
                    max_order = new_order

                    Address.objects.create(
                        project=project,
                        name=name,
                        latitude=coordinates['lat'],
                        longitude=coordinates['lon'],
                        order=new_order
                    )
                    num_processed += 1
                except ValueError as ve:
                    return JsonResponse({'status': 'error', 'message': str(ve), 'remaining_requests': counter.count})
        
        counter.count -= num_processed
        counter.save()
        
        return JsonResponse({'status': 'ok', 'remaining_requests': counter.count}, status=200)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

# удаление адреса из списка адресов в событии      
def delete_event_address(request, address_id):
    if request.method == "POST":
        address = get_object_or_404(EventAddress, id=address_id)
        address.delete()
        return JsonResponse({"success": True})
    else:
        return JsonResponse({"error": "Invalid request"}, status=400)

# сохранение новых адресов в событии    
@csrf_exempt
def add_new_addresses(request, event_id):
    reset_request_counter(request)
    counter = RequestCounter.objects.get(user=request.user)

    if counter.count <= 0:
        time_left = get_time_until_midnight()
        return JsonResponse({
            'status': 'error',
            'message': f'Закончились запросы для координат - запросы будут доступны через {time_left}',
            'remaining_requests': counter.count
        })

    try:
        event = get_object_or_404(Event, id=event_id)
        data = json.loads(request.body)
        new_addresses = data.get('new_addresses', [])
        num_processed = 0

        # Получаем наибольший order среди существующих адресов
        max_order = EventAddress.objects.filter(event=event).aggregate(Max('order'))['order__max'] or 0

        addresses_created = []
        for item in new_addresses:
            if item['name']:
                if counter.count <= 0:
                    time_left = get_time_until_midnight()
                    return JsonResponse({
                        'status': 'error',
                        'message': f'Закончились запросы для координат - запросы будут доступны через {time_left}',
                        'remaining_requests': counter.count
                    })

                coordinates = get_coordinates_from_yandex(request, item['name'])
                # Увеличьте max_order для нового адреса
                new_order = max_order + 1
                max_order = new_order  # Обновление для следующего адреса
                
                new_address = EventAddress.objects.create(
                    event=event,
                    name=item['name'],
                    latitude=coordinates['lat'],
                    longitude=coordinates['lon'],
                    order=new_order
                )
                addresses_created.append(new_address)
                num_processed += 1

        counter.count -= num_processed
        counter.save()

        response_data = {
            'status': 'ok',
            'remaining_requests': counter.count,
            'addresses': [
                {
                    'id': addr.id,
                    'name': addr.name,
                    'latitude': addr.latitude,
                    'longitude': addr.longitude
                }
                for addr in addresses_created
            ]
        }

        return JsonResponse(response_data)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

# обновление порядка адресов в событии
def update_address_order(request):
    if request.method == 'POST':
        order_data = json.loads(request.POST.get('order', '[]'))
        model_type = request.POST.get('model')

        if model_type == 'EventAddress':
            ModelClass = EventAddress
        elif model_type == 'Address':
            ModelClass = Address
        else:
            return JsonResponse({'status': 'failure', 'error': 'Invalid model type'}, status=400)

        for item in order_data:
            try:
                address = ModelClass.objects.get(id=item['id'])
                address.order = item['order']
                address.save()
            except ModelClass.DoesNotExist:
                return JsonResponse({'status': 'failure', 'error': 'Address not found'}, status=400)

        return JsonResponse({'status': 'success'})

    return JsonResponse({'status': 'failure'}, status=400)

@login_required
# представление списка назначенных событий для исполнителя
def assigned_events_list(request):
    current_user = request.user
    event_users_by_status = {
        'assigned': EventUser.objects.filter(user=current_user, status='assigned').order_by('event__event_date'),
        'confirmed': EventUser.objects.filter(user=current_user, status='confirmed').order_by('event__event_date'),
        'in_progress': EventUser.objects.filter(user=current_user, status='in_progress').order_by('event__event_date'),
        'completed': EventUser.objects.filter(user=current_user, status='completed').order_by('event__event_date'),
    }

    # Подсчёт количества адресов и преобразование статусов
    event_data_by_status = {}
    status_display_map = dict(EventUser.STATUS_CHOICES)

    manager_api_key = None

    for status, events in event_users_by_status.items():
        status_display = status_display_map.get(status, status)
        event_data_by_status[status] = {
            'display': status_display,
            'count': events.count(),
            'events': [
                (e.event, e.event.addresses.filter(assigned_user=current_user).count())
                for e in events
            ]
        }

        # Получаем API-ключ менеджера для первых доступных событий
        if manager_api_key is None and events.exists():
            # Предполагаем, что `event.project` ссылается на проект с полем `user`
            first_event = events.first().event
            # Доступ к менеджеру проекта: `first_event.project.user`
            manager_api_key = first_event.project.user.profile.api_key

    # Для статуса 'in_progress', объединяем адреса
    in_progress_addresses = []
    in_progress_count = 0
    if 'in_progress' in event_users_by_status:
        in_progress_events = event_users_by_status['in_progress']
        in_progress_count = in_progress_events.count()  # Получаем количество событий 'in_progress'
        for event_user in in_progress_events:
            event_addresses = event_user.event.addresses.filter(assigned_user=current_user)
            in_progress_addresses.extend(event_addresses)

    return render(request, 'projects/assigned_events_list.html', {
        'event_data_by_status': event_data_by_status,
        'in_progress_addresses': in_progress_addresses if in_progress_addresses else None,
        'manager_api_key': manager_api_key,  # Передаем API ключ менеджера
        'in_progress_count': in_progress_count,  # Передаем count в шаблон
    })

@login_required
# представление для удаления назначенного события из списка назначенных событий
def remove_assigned_event(request, event_id):
    event = get_object_or_404(Event, id=event_id)

    if request.method == 'POST':
        # Изменяем статус на "отказ"
        EventUser.objects.filter(event=event, user=request.user).update(status='declined')
        # Убираем назначение пользователя
        EventAddress.objects.filter(event=event, assigned_user=request.user).update(assigned_user=None)
        return redirect('assigned_events_list')

    return redirect('assigned_events_list')

@csrf_exempt
# сохраняет и динамически обновляет исполнителей события
def update_event_executors(request, event_id):
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        executors_ids = request.POST.getlist('executors[]')
        event = get_object_or_404(Event, pk=event_id)

        # Снять всех текущих исполнителей
        # event.assigned_users.clear()

        # Назначить новых исполнителей
        users = User.objects.filter(id__in=executors_ids)
        event.assigned_users.add(*users)

        executors_data = [{'id': user.id,
                           'name': user.executorprofile.name,
                           'district': user.executorprofile.district,
                           'phone_number': user.executorprofile.phone_number}
                          for user in users]

        return JsonResponse({'executors': executors_data})

    return JsonResponse({'error': 'Invalid request'}, status=400)

# Удаляем пользователя из исполнителей события и отвязываем адреса
def remove_executor(request):
    if request.method == 'POST':
        executor_id = request.POST.get('executor_id')
        event_id = request.POST.get('event_id')

        try:
            event = Event.objects.get(id=event_id)
            user = User.objects.get(id=executor_id)
        except (Event.DoesNotExist, User.DoesNotExist):
            return JsonResponse({'error': 'Event or User not found'}, status=404)

        # Удаляем пользователя из исполнителей события
        event.assigned_users.remove(user)

        # Находим все EventAddress, связанные с этим событием и этим пользователем, и отвязываем пользователя
        EventAddress.objects.filter(event=event, assigned_user=user).update(assigned_user=None)

        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Invalid request'}, status=400)

@csrf_exempt
@require_POST
# назначение исполнителя
def assign_executor(request):
    event_id = request.POST.get('event_id')
    user_id = request.POST.get('user_id')
    address_indexes = request.POST.get('address_indexes')

    if not event_id or not user_id:
        return HttpResponseBadRequest("Invalid data")

    try:
        event = Event.objects.get(id=event_id)
        user = User.objects.get(id=user_id)

        event_user, created = EventUser.objects.get_or_create(event=event, user=user)

        # Преобразуем полученные адреса в список целых чисел
        address_indexes = json.loads(address_indexes) if address_indexes else []

        if any(not isinstance(i, int) for i in address_indexes):
            return JsonResponse({"success": False, "message": "Введён некорректный номер адреса."})

        # Используем количество адресов, принадлежащих этому событию
        total_event_addresses = event.addresses.count()

        # Добавляем проверку для положительных индексов
        if any(index <= 0 for index in address_indexes):
            return JsonResponse({"success": False, "message": f"Введите номера в диапазоне от 1 до {total_event_addresses}."})

        if any(index > total_event_addresses for index in address_indexes):
            return JsonResponse({"success": False, "message": f"Введите номера в существующем диапазоне от 1 до {total_event_addresses}."})

        restricted_statuses = ['confirmed', 'in_progress', 'completed']
        selected_addresses = []
        already_assigned = True
        event_addresses = list(event.addresses.all())  # Получаем адреса для события

        for index in address_indexes:
            if 0 < index <= len(event_addresses):
                address = event_addresses[index-1]
                if (address.assigned_user and
                    address.assigned_user != user and
                    EventUser.objects.filter(user=address.assigned_user, event=event, status__in=restricted_statuses).exists()):
                    return JsonResponse({"success": False, "message": "Некоторые адреса уже назначены другому исполнителю."})

                if address.assigned_user != user:
                    already_assigned = False

                selected_addresses.append(address)

        # Если все выбранные адреса уже назначены данному пользователю
        if already_assigned:
            return JsonResponse({"success": False, "message": "Введённые адреса уже назначены данному исполнителю."})

        # Назначение и изменение статуса
        for address in selected_addresses:
            address.assigned_user = user
            address.save()

        if event_user.status == 'completed' and selected_addresses:
            event_user.status = 'in_progress'
        elif event_user.status not in ['confirmed', 'in_progress']:
            event_user.status = 'assigned' if selected_addresses else 'chosen'

        event_user.save()

        return JsonResponse({"success": True})

    except Event.DoesNotExist:
        return HttpResponseBadRequest("Событие не найдено")
    except User.DoesNotExist:
        return HttpResponseBadRequest("Пользователь не найден")
    except json.JSONDecodeError:
        return HttpResponseBadRequest("Некорректный формат данных")
    
@login_required
@require_POST
# функция для обработки статуса подтверждения
def confirm_executor(request):
    event_id = request.POST.get('event_id')
    user_id = request.POST.get('user_id')

    try:
        event_user = EventUser.objects.get(event_id=event_id, user_id=user_id)
        event_user.status = 'confirmed'  # Устанавливаем статус на "подтверждён"
        event_user.save()
        return JsonResponse({"success": True})
    except EventUser.DoesNotExist:
        return JsonResponse({"error": "EventUser not found"}, status=404)
    
# функция для обработки статуса завершения
@require_POST
@login_required
def complete_event(request):
    event_id = request.POST.get('event_id')
    executor_profile = getattr(request.user, 'executorprofile', None)

    if not executor_profile:
        return JsonResponse({'success': False, 'error': 'Вы не являетесь исполнителем.'}, status=400)

    event = get_object_or_404(Event, pk=event_id)
    assigned_event_addresses = event.addresses.filter(assigned_user=request.user)
    addresses_without_photos = []

    for address in assigned_event_addresses:
        executor_id = executor_profile.id
        project_user_id = event.project.user.id
        project_id = str(event.project.id)
        event_type = event.get_event_type_display()
        address_name = re.sub(r'[\\/*?:"<>|]', "_", address.name)

        base_path = os.path.join(
            settings.MEDIA_ROOT, str(project_user_id), project_id, event_type, str(executor_id)
        )
        problems_path = os.path.join(base_path, "problems")

        has_photos = False
        for path in [base_path, problems_path]:
            if os.path.exists(path):
                for file_name in os.listdir(path):
                    if file_name.startswith(address_name):
                        has_photos = True
                        break

        if not has_photos:
            addresses_without_photos.append(address.name)

    if addresses_without_photos:
        return JsonResponse({'success': False, 'error_addresses': addresses_without_photos}, status=200)

    # Update status to 'completed'
    event_user = EventUser.objects.get(user=request.user, event=event)
    event_user.status = 'completed'
    event_user.save()

    return JsonResponse({'success': True}, status=200)

# для смены статуса назначенного и подтверждённого события на "в работе"
def start_event(request):
    if request.method == 'POST':
        event_id = request.POST.get('event_id')
        user = request.user

        try:
            event_user = EventUser.objects.get(event_id=event_id, user=user)

            # Превращаем дату события в объект datetime и делаем её aware
            event_start_date_naive = datetime.combine(event_user.event.event_date, time.min)
            event_start_date = make_aware(event_start_date_naive)

            current_time = timezone.now()

            # Проверка: можем ли мы начать событие
            if (event_start_date - current_time).days <= 1:
                event_user.status = 'in_progress'
                event_user.save()
                return JsonResponse({"success": True})
            else:
                return JsonResponse({"success": False, "error": "Событие можно будет начать не раньше, чем за день до запланированной даты"})

        except EventUser.DoesNotExist:
            return JsonResponse({"error": "EventUser not found"}, status=404)
        except Exception as e:
            # Логирование исключений для отслеживания
            print(f"Exception occurred: {str(e)}")
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Некорректный запрос"}, status=400)

from collections import defaultdict
# назначенные адреса для event_detail.html
@login_required
def event_detail(request, event_id):
    event = get_object_or_404(Event, pk=event_id)
    current_user = request.user
    is_manager = current_user.groups.filter(name='Менеджер').exists()
    is_executor = current_user.groups.filter(name='Исполнитель').exists()
    manager_addresses = defaultdict(list)
    executor_addresses = []

    def process_addresses(address_list):
        addresses_with_photos = []
        executor_colors = {'not_assigned': '#808080'}
        color_palette = ['#FF5733', '#33FF57', '#3357FF', '#F3FF33', '#FF33A6']
        color_index = 0
        for address in address_list:
            executor_id = address.assigned_user.executorprofile.id if address.assigned_user else 'not_assigned'
            project_user_id = event.project.user.id
            project_id = str(event.project.id)
            event_type = event.get_event_type_display()
            address_name = re.sub(r'[\\/*?:"<>|]', "_", address.name)
            base_path = os.path.join(settings.MEDIA_ROOT, str(project_user_id), project_id, event_type, str(executor_id))
            problems_path = os.path.join(base_path, "problems")
            has_photos = False
            if executor_id and os.path.exists(base_path):
                for file_name in os.listdir(base_path):
                    if file_name.startswith(address_name):
                        has_photos = True
                        break
            if executor_id and os.path.exists(problems_path):
                for file_name in os.listdir(problems_path):
                    if file_name.startswith(address_name):
                        has_photos = True
                        break
            if executor_id not in executor_colors:
                executor_colors[executor_id] = color_palette[color_index % len(color_palette)]
                color_index += 1
            addresses_with_photos.append({
                'address': address,
                'has_photos': has_photos,
                'executor_id': executor_id,
                'executor_color': executor_colors[executor_id]
            })
        return addresses_with_photos

    if is_manager and event.project.user == current_user:
        for address in event.addresses.all().select_related('assigned_user'):
            executor = address.assigned_user.executorprofile if address.assigned_user else None
            manager_addresses[executor].append(address)
        addresses_with_photos_manager = {executor: process_addresses(addresses) for executor, addresses in manager_addresses.items()}

        # Add unassigned addresses as a category "None"
        unassigned_addresses = manager_addresses.pop(None, [])
        addresses_with_photos_manager[None] = process_addresses(unassigned_addresses)
    else:
        addresses_with_photos_manager = {}

    if is_executor:
        executor_addresses = event.addresses.filter(assigned_user=current_user)
    addresses_with_photos_executor = process_addresses(executor_addresses)

    user_event_status = None
    if is_executor:
        try:
            event_user = EventUser.objects.get(user=current_user, event=event)
            user_event_status = event_user.status
        except EventUser.DoesNotExist:
            user_event_status = None

    return render(request, 'projects/event_detail.html', {
        'event': event,
        'addresses_with_photos_manager': addresses_with_photos_manager,
        'addresses_with_photos_executor': addresses_with_photos_executor,
        'is_executor': is_executor,
        'is_manager': is_manager,
        'user_event_status': user_event_status,
    })

# Оптимальный маршрут для event_detail.html
@csrf_exempt
def calculate_optimal_route(request):
    if request.method == 'POST':
        data = json.loads(request.POST['coordinates'])
        optimal_route = nearest_neighbor_event_detail(data)
        return JsonResponse({'route': optimal_route})

    return JsonResponse({'error': 'Invalid request method.'}, status=400)

@login_required
def events_control(request):
    current_user = request.user
    
    # Получаем все события, где проекты принадлежат текущему пользователю, с проверкой на статус события
    event_users_by_status = {
        'assigned': EventUser.objects.filter(event__project__user=current_user, status='assigned').order_by('event__event_date'),
        'confirmed': EventUser.objects.filter(event__project__user=current_user, status='confirmed').order_by('event__event_date'),
        'in_progress': EventUser.objects.filter(event__project__user=current_user, status='in_progress').order_by('event__event_date'),
        'completed': EventUser.objects.filter(event__project__user=current_user, status='completed').order_by('event__event_date'),
    }
    
    # Подсчёт количества адресов и преобразование статусов
    event_data_by_status = {}
    status_display_map = dict(EventUser.STATUS_CHOICES)
    
    for status, events in event_users_by_status.items():
        # Получаем отображаемое значение статуса
        status_display = status_display_map.get(status, status)
        event_data_by_status[status] = {
            'display': status_display,
            'count': events.count(),
            'events': [
                (
                    e.event,
                    e.event.addresses.filter(assigned_user=e.user).count(),
                    e.user  # добавляем пользователя в кортеж, чтобы показать исполнителя
                )
                for e in events.select_related('user', 'event__project')  # оптимизируем запрос
            ]
        }

    return render(request, 'projects/events_control.html', {
        'event_data_by_status': event_data_by_status,
        'current_user': current_user,
    })

# обновление счётчика запросов в шаблоне
def get_remaining_requests(request):
    user = request.user
    counter, created = RequestCounter.objects.get_or_create(user=user)
    reset_request_counter(request)  # Обновляем состояние счетчика
    return JsonResponse({'remaining_requests': counter.count})

@login_required
def upload_photos(request, event_id, address_id):
    if request.method == 'POST' and request.user.groups.filter(name='Исполнитель').exists():
        event = get_object_or_404(Event, pk=event_id)
        address = get_object_or_404(EventAddress, pk=address_id, event=event)
        photos = request.FILES.getlist('photos')
        force_mjeure = request.POST.get('force_mjeure') == "on"

        # Проверка на количество фото только если не форс-мажор
        if not force_mjeure and len(photos) != event.photo_count:
            return JsonResponse({
                'error': f'Количество фотографий должно быть равно {event.photo_count}.'
            }, status=200)

        executor_id = request.user.executorprofile.id
        project_user_id = event.project.user.id
        project_id = str(event.project.id)
        event_type = event.get_event_type_display()
        address_name = re.sub(r'[\\/*?:"<>|]', "_", address.name)

        # Определите пути для главной и "проблемной" папок
        base_path = os.path.join(settings.MEDIA_ROOT, str(project_user_id), project_id, event_type, str(executor_id))
        problems_path = os.path.join(base_path, "problems")

        # Удалить все существующие фотографии для этого адреса из обеих папок
        for path in [base_path, problems_path]:
            if os.path.exists(path):
                for file_name in os.listdir(path):
                    if file_name.startswith(address_name):
                        file_path = os.path.join(path, file_name)
                        default_storage.delete(file_path)

        # Установление пути сохранения в зависимости от флага "Форс-мажор"
        if force_mjeure:
            base_path = problems_path
            
        # Создать дирректорию если не существует
        os.makedirs(base_path, exist_ok=True)

        # Сохранение новых фотографий
        def generate_unique_file_path(base_path, base_name, extension):
            counter = 0
            unique_file_path = os.path.join(base_path, f"{base_name}{extension}")

            while default_storage.exists(unique_file_path):
                counter += 1
                unique_file_path = os.path.join(base_path, f"{base_name}_{counter}{extension}")

            return unique_file_path

        for photo in photos:
            file_extension = os.path.splitext(photo.name)[1]
            unique_file_path = generate_unique_file_path(base_path, address_name, file_extension)

            with default_storage.open(unique_file_path, 'wb+') as destination:
                for chunk in photo.chunks():
                    destination.write(chunk)

        return JsonResponse({'success': 'Фотографии успешно обновлены'}, status=200)

    return JsonResponse({'error': 'Invalid request'}, status=400)

@login_required
def view_photos(request, event_id):
    if request.method == 'GET' and request.user.groups.filter(name='Менеджер').exists():
        event = get_object_or_404(Event, pk=event_id)
        address_id = request.GET.get('address_id')
        address = get_object_or_404(EventAddress, pk=address_id, event=event)

        executor_id = address.assigned_user.executorprofile.id if address.assigned_user else None
        project_user_id = event.project.user.id
        project_id = str(event.project.id)
        event_type = event.get_event_type_display()
        address_name = re.sub(r'[\\/*?:"<>|]', "_", address.name)

        # Проверка обеих директорий: обычной и проблемной
        photos = []
        for subfolder in ["", "problems"]:
            base_path = os.path.join(settings.MEDIA_ROOT, str(project_user_id), project_id, event_type, str(executor_id), subfolder)

            if os.path.exists(base_path):
                for file_name in os.listdir(base_path):
                    if file_name.startswith(address_name):
                        file_path = os.path.join(base_path, file_name)
                        url = default_storage.url(os.path.relpath(file_path, settings.MEDIA_ROOT))
                        photos.append({
                            'url': url,
                            'name': file_name,
                            'address_name': address_name
                        })
        
        return JsonResponse({'photos': photos})

    return JsonResponse({'error': 'Invalid request'}, status=400)

@login_required
def fetch_executor_photos(request):
    user_id = request.GET.get('user_id')
    event_id = request.GET.get('event_id')

    event = get_object_or_404(Event, id=event_id)
    user = get_object_or_404(User, id=user_id)
    event_user = get_object_or_404(EventUser, user=user, event=event)

    executor_id = user.executorprofile.id
    base_path = os.path.join(settings.MEDIA_ROOT, str(event.project.user.id), str(event.project.id), event.get_event_type_display(), str(executor_id))
    problems_path = os.path.join(base_path, "problems")

    photos = []
    problems = []

    def load_and_filter_photos(directory_path, assigned_names=None, is_problem=False):
        if assigned_names is None:
            assigned_names = []

        if os.path.exists(directory_path):
            for file_name in os.listdir(directory_path):
                file_path = os.path.join(directory_path, file_name)
                if os.path.isfile(file_path):
                    photo_url = default_storage.url(file_path)
                    # Добавляем фото, если:
                    # - (статус "completed" и имя начинается с одного из assigned_names), или
                    # - (статус не "completed" и нам нужно загрузить все)
                    if (event_user.status == 'completed' and any(file_name.startswith(addr_name) for addr_name in assigned_names)) or event_user.status != 'completed':
                        if is_problem:
                            problems.append({'url': photo_url, 'name': file_name})
                        else:
                            photos.append({'url': photo_url, 'name': file_name})
                    elif event_user.status == 'completed':
                        # Удаляем файл, если статус "completed" и он не соответствует именам
                        os.remove(file_path)

    if event_user.status == 'completed':
        assigned_addresses = event.addresses.filter(assigned_user=user)
        assigned_address_names = [re.sub(r'[\\/*?:"<>|]', "_", address.name) for address in assigned_addresses]

        # Загружаем и фильтруем файлы. Удаляем только при статусе "completed".
        removed_photos = load_and_filter_photos(base_path, assigned_address_names)
        removed_problems = load_and_filter_photos(problems_path, assigned_address_names, is_problem=True)

    else:
        # Просто загружаем все фото, без фильтрации и удаления
        load_and_filter_photos(base_path, [])
        load_and_filter_photos(problems_path, [], is_problem=True)

    # Вернуть данные о фотографиях
    return JsonResponse({'photos': photos, 'problems': problems})

def download_executor_photos(request):
    user_id = request.GET.get('user_id')
    event_id = request.GET.get('event_id')

    event = get_object_or_404(Event, id=event_id)
    user = get_object_or_404(User, id=user_id)
    event_user = get_object_or_404(EventUser, user=user, event=event)
    executor_id = user.executorprofile.id

    # Формируем название архива
    project_name = event.project.name
    event_type = event.get_event_type_display()
    executor_name = user.executorprofile.name
    zip_filename = f"{project_name} - {event_type} - {executor_name}.zip"

    # Определяем пути к фотографиям
    base_path = os.path.join(settings.MEDIA_ROOT, str(event.project.user.id), str(event.project.id),
                             event.get_event_type_display(), str(executor_id))
    problems_path = os.path.join(base_path, "problems")
    photos = []

    # Собираем все фотографии из основной директории
    if os.path.exists(base_path):
        for file_name in os.listdir(base_path):
            file_path = os.path.join(base_path, file_name)
            if os.path.isfile(file_path):
                photos.append((file_path, file_name))

    # Собираем все фотографии из директории "problems" при её существовании
    if os.path.exists(problems_path):
        for file_name in os.listdir(problems_path):
            file_path = os.path.join(problems_path, file_name)
            if os.path.isfile(file_path):
                photos.append((file_path, os.path.join("problems", file_name)))

    # Создаем архив с фотографиями
    if photos:
        resp = HttpResponse(content_type='application/zip')
        # Используем iri_to_uri для экранирования возможно неподдерживаемых символов в HTTP заголовках
        resp['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{iri_to_uri(zip_filename)}'

        with zipfile.ZipFile(resp, 'w') as zf:
            for file_path, arc_name in photos:
                zf.write(file_path, arc_name)

        return resp
    else:
        return HttpResponse("No photos available for download.", status=404)
    
@login_required
@csrf_exempt
def get_addresses_for_events(request):
    if request.method == 'POST':
        try:
            event_ids = json.loads(request.POST.get('event_ids', '[]'))
            addresses = EventAddress.objects.filter(
                event__in=event_ids,
                assigned_user=request.user
            )
            
            address_list = [
                {
                    'id': address.id,
                    'name': address.name,
                    'latitude': address.latitude,
                    'longitude': address.longitude,
                    'projectName': address.event.project.name,
                    'eventTypeDisplay': address.event.get_event_type_display(),
                    'photos_uploaded': check_photos_uploaded(address, request.user)  # Проверяем наличие фото
                }
                for address in addresses
            ]
            
            return JsonResponse({'success': True, 'addresses': address_list})
        
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Некорректный формат JSON'})
        
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
@csrf_exempt
# объединение адресов со статусом "в работе" для assigned_events_list.html
def save_combined_address_order(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.POST.get('order', '[]'))

            for index, address_id in enumerate(data):
                if not address_id:
                    return JsonResponse({'success': False, 'error': 'Некорректный идентификатор адреса.'})

                try:
                    address_id = int(address_id)
                except ValueError:
                    return JsonResponse({'success': False, 'error': 'Идентификатор адреса должен быть числом.'})

                try:
                    address = EventAddress.objects.get(id=address_id, assigned_user=request.user)
                    address.order = index  # Устанавливаем новый порядок
                    address.save()
                except EventAddress.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'Не найден адрес или нет доступа.'})

            return JsonResponse({'success': True})
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Некорректный формат JSON'})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})

# логика оптимизации маршрута для assigned_events_list.html
def optimize_route(request):
    if request.method == 'POST':
        try:
            coords = json.loads(request.POST.get('coords', '[]'))
            print('Received coords:', coords)  # Для отладки

            # Конвертируем координаты в формат [(lat, lon), ...]
            coord_tuples = [(coord['latitude'], coord['longitude']) for coord in coords]

            # Используем функцию ближайшего соседа для оптимизации порядка
            order = nearest_neighbor(coord_tuples)

            # Создаем список упорядоченных ID адресов
            ordered_ids = [coords[i]['id'] for i in order]

            return JsonResponse({'success': True, 'ordered_ids': ordered_ids})
        except KeyError as e:
            return JsonResponse({'success': False, 'error': f'Key error: {e}'})
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Некорректный формат JSON'})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

from django.core.files.storage import FileSystemStorage

def upload_combined_addresses_photos(request):
    if request.method == 'POST':
        address_id = request.POST.get('address_id')
        force_majeure = request.POST.get('force_majeure') == 'on'
        photos = request.FILES.getlist('photos')

        try:
            address = EventAddress.objects.get(id=address_id, assigned_user=request.user)
            event = address.event
            event_type_display = re.sub(r'[\\/*?:"<>|]', "_", event.get_event_type_display())
            address_name_safe = re.sub(r'[\\/*?:"<>|]', "_", address.name)

            # Основная директория для фотографий
            base_folder = os.path.join(
                str(event.project.user.id), 
                str(event.project.id), 
                event_type_display, 
                str(request.user.executorprofile.id)
            )

            # Директория для проблемных фотографий
            problems_folder = os.path.join(base_folder, 'problems')

            # Определяем путь фотодиректории, в зависимости от форс-мажора
            if force_majeure:
                photo_dir = os.path.join(settings.MEDIA_ROOT, problems_folder)
            else:
                photo_dir = os.path.join(settings.MEDIA_ROOT, base_folder)

            # Удаляем все фотографии как в основной, так и в проблемной директориях
            def clear_photos(directory):
                if os.path.exists(directory):
                    for file in os.listdir(directory):
                        if file.startswith(address_name_safe):
                            os.remove(os.path.join(directory, file))

            # Удаляем старые фотографии перед загрузкой новых
            clear_photos(os.path.join(settings.MEDIA_ROOT, base_folder))
            clear_photos(os.path.join(settings.MEDIA_ROOT, problems_folder))

            # Проверка количества загружаемых фото
            if force_majeure:
                # Если форс-мажор, проверяем количество фото
                if not (1 <= len(photos) <= 10):
                    return JsonResponse({
                        'success': False,
                        'error': 'Вы должны загрузить от 1 до 10 фотографий'
                    })
            else:
                # Проверяем соответствие требуемому количеству фото
                required_photos = event.photo_count
                if len(photos) != required_photos:
                    return JsonResponse({
                        'success': False,
                        'error': f'Кол-во загружаемых фото должно быть - {required_photos}'
                    })

            # Создаем директорию для сохранения новых фотографий, если она не существует
            os.makedirs(photo_dir, exist_ok=True)

            # Сохранение новых файлов
            fs = FileSystemStorage(location=photo_dir)
            for i, photo in enumerate(photos):
                # Добавляем суффикс начиная со второй фотографии
                name_suffix = f"_{i}" if i > 0 else ""
                fs.save(f"{address_name_safe}{name_suffix}.jpg", photo)

            return JsonResponse({'success': True})

        except EventAddress.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Неверный адрес'})

    return JsonResponse({'success': False, 'error': 'Неверный метод запроса'})

# Функция для проверки загруженных фото у совмещённого списка адресов на assigned_events_list.html
def check_photos_uploaded(address, user):
    event = address.event
    event_type_display = re.sub(r'[\\/*?:"<>|]', "_", event.get_event_type_display())
    address_name_safe = re.sub(r'[\\/*?:"<>|]', "_", address.name)

    # Основная папка и папка "problems"
    base_folder = os.path.join(
        str(event.project.user.id),
        str(event.project.id),
        event_type_display,
        str(user.executorprofile.id)
    )
    problems_folder = os.path.join(base_folder, 'problems')

    # Проверка существования фото в обеих папках
    return (
        check_photos_in_directory(os.path.join(settings.MEDIA_ROOT, base_folder), address_name_safe) or
        check_photos_in_directory(os.path.join(settings.MEDIA_ROOT, problems_folder), address_name_safe)
    )

def check_photos_in_directory(photo_dir, address_name_safe):
    if not os.path.exists(photo_dir):
        return False
    
    return any(photo.startswith(address_name_safe) for photo in os.listdir(photo_dir))